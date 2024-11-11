from flask import Flask, render_template, request, jsonify
# import whisper
import os
# from openai import OpenAI
import json
from docx import Document
import openpyxl
import io

class TranscriptionApp:
    def __init__(self):
        self.app = Flask(__name__)
        # self.model = whisper.load_model("small")  # モデル指定

        self.setup_routes()

    def setup_routes(self):
        self.app.add_url_rule('/', 'index', self.index)
        self.app.add_url_rule('/transcribe', 'transcribe', self.transcribe, methods=['POST'])
        self.app.add_url_rule('/create_minutes', 'create_minutes', self.create_minutes, methods=['POST'])  # 追加
        self.app.add_url_rule('/prompt', 'get_prompt', self.get_prompt, methods=['GET'])
        self.app.add_url_rule('/favicon.ico', 'favicon', self.favicon)

    def index(self):
        return render_template('index.html')

    def transcribe(self):
        if 'file' not in request.files:
            return jsonify({'error': 'No file part'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No selected file'}), 400

        # 一時ファイルとして保存
        file_path = os.path.join('temp', file.filename)
        file.save(file_path)

        # 音声ファイルの転写
        # result = self.model.transcribe(file_path, verbose=True, fp16=False, language="ja")
        # transcribed_text = result['text']  # 生データ（文字おこし結果）

        # 一時ファイルを削除
        os.remove(file_path)

        # クライアント側から送られたプロンプトデータを使用する
        isAdvancedModeChecked = request.form.get('isAdvancedModeChecked')

        if isAdvancedModeChecked=='true':
            prompt = request.form.get('prompt')
        else:
            participants = json.loads(request.form.get('participants'))
            meetingDate = request.form.get('meetingDate')
            if 'fileFormatData' in request.files:
                fileFormatData = request.files['fileFormatData']
                filename = fileFormatData.filename
                fileFormatDataStr = self.extract_text(fileFormatData, filename)
            else :
                fileFormatDataStr = ''
            prompt = self.set_script_beginner(participants, meetingDate, fileFormatDataStr)
        # text = f'{prompt} {transcribed_text}'

        # messages = [{"role": "system", "content": text}]

        # completion = self.client.chat.completions.create(
        #     model="gpt-4o-mini",
        #     messages=messages
        # )

        # JSONで生データと要約データを一緒に返す
        return jsonify({
            'transcription': f'議事録・文字起こし{prompt}',
            'raw_transcription': f'議事録・文字起こし{prompt}'
        })

    def create_minutes(self):
        # クライアントからの文字おこしデータを取得
        transcription = request.form.get('transcription')
        isAdvancedModeChecked = request.form.get('isAdvancedModeChecked')

        if not transcription:
            return jsonify({'error': 'No transcription data provided'}), 400

        if isAdvancedModeChecked=='true':
            prompt = request.form.get('prompt')
        else:
            participants = json.loads(request.form.get('participants'))
            meetingDate = request.form.get('meetingDate')
            if 'fileFormatData' in request.files:
                fileFormatData = request.files['fileFormatData']
                filename = fileFormatData.filename
                fileFormatDataStr = self.extract_text(fileFormatData, filename)
            else :
                fileFormatDataStr = ''
            prompt = self.set_script_beginner(participants, meetingDate, fileFormatDataStr)
        # 議事録作成用のプロンプトを生成
        # text = f'{prompt} {transcription}'

        # messages = [{"role": "system", "content": f"{text}"}]

        # completion = self.client.chat.completions.create(
        #     model="gpt-4o-mini",
        #     messages=messages
        # )

        # 議事録を生成して返す
        return jsonify({
            'transcription': f'議事録のみ{prompt}'
        })

    def get_prompt(self):
        # prompt.txtファイルの内容を読み込んで返す
        try:
            with open('prompt.txt', 'r', encoding='utf-8') as file:
                prompt_content = file.read()
            return jsonify({'prompt': prompt_content})
        except FileNotFoundError:
            return jsonify({'error': 'prompt.txtファイルが見つかりませんでした。'}), 404

    def favicon(self):
        return '', 204  # No Content
    
    def set_script_beginner(self, participants, meetingDate, fileFormatData):
        # スクリプトをファイルから読み込む
        with open('prompt_beginner.txt', 'r', encoding='utf-8') as file:
            script = file.read()

        if meetingDate:
            result_meetingDate = meetingDate
        else:
            result_meetingDate = '「未記入」と記載してください'

        if any(participants):  # 配列に一つでも値があればTrue
            result_participants = ",".join([name for name in participants if name])
        else:
            result_participants = '「未記入」と記載してください'

        # プレースホルダーの置き換え
        script = script.format(
            meetingDate = result_meetingDate,
            participants = result_participants,
            fileFormatData = fileFormatData,
        )
        return script
        
    def extract_text_from_txt(self, file_data):
        """テキストファイルから文字列を抽出する関数"""
        content = file_data.read().decode('utf-8')
        return content

    def extract_text_from_docx(self, file_data):
        """Wordファイルから文字列を抽出する関数"""
        doc = Document(file_data)
        content = [paragraph.text for paragraph in doc.paragraphs]
        return '\n'.join(content)

    def extract_text_from_excel(self, file_data):
        """Excelファイルから文字列を抽出する関数"""
        try:
            file_data.seek(0)
            file_data = io.BytesIO(file_data.read())
            wb = openpyxl.load_workbook(file_data, data_only=True)
            content = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = [str(cell) if cell is not None else '' for cell in row]
                    content.append('\t'.join(row_text))
            return '\n'.join(content)
        except Exception as e:
            return f"Excelファイルの読み取り中にエラーが発生しました: {e}"

    def extract_text(self, file_data, filename):
        """ファイル形式に応じて適切な関数を呼び出して文字列を抽出する関数"""
        file_data.seek(0)  # ファイルポインタを先頭に戻す
        if filename.endswith('.txt'):
            return self.extract_text_from_txt(file_data)
        elif filename.endswith('.docx'):
            return self.extract_text_from_docx(file_data)
        elif filename.endswith('.xlsx'):
            return self.extract_text_from_excel(file_data)

# クラスの外でサーバーを立ち上げる処理
transcription_app = TranscriptionApp()
app = transcription_app.app

if __name__ == '__main__':
    os.makedirs('temp', exist_ok=True)
    app.run(debug=True, host='0.0.0.0', port=5001)